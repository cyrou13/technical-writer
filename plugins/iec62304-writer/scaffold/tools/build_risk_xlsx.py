#!/usr/bin/env python3
"""Build the QMS-ready 4-sheet Excel risk inventory.

Equivalent of the Avicenna AV-DP-CINA-CSP-10-005-annex1-RISK-TABLE.xlsx —
one sheet per risk category (Design, Production, Usability, Cybersecurity)
with one row per item and full ISO 14971 + IEC 81001-5-1 columns.

Reads:
    dt-config.yaml                       (QMS metadata: cover identifier)
    docs/items/RSK/*.md                  → "Design risk analysis" sheet
    docs/items/PRSK/*.md                 → "Production risk analysis" sheet
    docs/items/URSK/*.md                 → "Usability risk analysis" sheet
    docs/items/THR/*.md                  → "Cybersecurity risk analysis" sheet
    docs/items/{SRS,SDS,TC}/*.md         (optional — to enumerate mitigators)

Writes:
    docs/export/<identifier>-<version_label>-RISK-TABLE.xlsx
    docs/export/<identifier>-<version_label>-risk-xlsx.log

Requires the third-party `openpyxl` package. Stdlib alone cannot produce a
valid .xlsx without re-implementing the OOXML zip+xml format by hand. If
`openpyxl` is missing, the script exits 1 with a clear install hint.

Python 3.12+.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Shared helpers — see tools/_lib.py
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _lib import (  # noqa: E402
    Item,
    load_items,
    parse_yaml,
    risk_index,
)

ROOT = Path.cwd()
CONFIG_PATH = ROOT / "dt-config.yaml"
ITEMS_DIR = ROOT / "docs" / "items"
EXPORT_DIR = ROOT / "docs" / "export"




# ---------------------------------------------------------------------------
# Build context + styling
# ---------------------------------------------------------------------------


@dataclass
class BuildContext:
    config: dict
    rsk: list[Item]
    prsk: list[Item]
    ursk: list[Item]
    thr: list[Item]
    mitigators: list[Item]
    log_lines: list[str] = field(default_factory=list)
    red_cells: int = 0

    def log(self, msg: str) -> None:
        self.log_lines.append(msg)
        print(msg, file=sys.stderr)

    def controls_for(self, risk_id: str) -> list[Item]:
        return [m for m in self.mitigators if risk_id in m.mitigates]


# Styles
HEADER_FONT = None
HEADER_FILL = None
HEADER_ALIGN = None
RED_FILL = None
WRAP_ALIGN = None


def init_styles() -> None:
    """Initialize style globals once openpyxl is confirmed available."""
    global HEADER_FONT, HEADER_FILL, HEADER_ALIGN, RED_FILL, WRAP_ALIGN
    HEADER_FONT = Font(bold=True, color="000000")
    HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)


def apply_headers(ws, headers: list[str], widths: list[int]) -> None:
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32


def write_row(ws, row_idx: int, values: list, residual_acceptable_col: int | None, ctx: BuildContext) -> None:
    """Write a data row. If `residual_acceptable_col` is provided and the value
    at that column index is False (or "False"), highlight the row's residual-
    acceptable cell red."""
    res_val = values[residual_acceptable_col - 1] if residual_acceptable_col else None
    is_red = res_val in (False, "False", "false")
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
        cell.alignment = WRAP_ALIGN
        if is_red and col_idx == residual_acceptable_col:
            cell.fill = RED_FILL
            ctx.red_cells += 1


# ---------------------------------------------------------------------------
# Sheet 1: Design risk analysis (RSK)
# ---------------------------------------------------------------------------


DESIGN_HEADERS = [
    "Risk ID", "Software Function", "Software Item", "Hazard",
    "Initiating causes", "Foreseeable sequence of events",
    "Hazardous situation", "Harm",
    "Initial Probability", "Initial Severity", "Initial Risk Index", "Initial Risk Level", "Initial Acceptable",
    "Control Hierarchy", "Mitigating Items",
    "Residual Probability", "Residual Severity", "Residual Risk Index", "Residual Risk Level", "Residual Acceptable",
    "Arising Risks", "Labeling Disclosure",
]
DESIGN_WIDTHS = [16, 22, 22, 32, 40, 40, 32, 32, 14, 14, 12, 14, 14, 22, 22, 14, 14, 12, 14, 14, 22, 30]


def build_design_sheet(wb, ctx: BuildContext) -> None:
    ws = wb.create_sheet("Design risk analysis")
    apply_headers(ws, DESIGN_HEADERS, DESIGN_WIDTHS)
    active = [r for r in ctx.rsk if r.status != "Deprecated"]
    for row_idx, r in enumerate(sorted(active, key=lambda i: i.id), start=2):
        sev, prob = r.get("severity"), r.get("probability")
        rsev, rprob = r.get("residual_severity"), r.get("residual_probability")
        controls = ctx.controls_for(r.id)
        arising = r.get("arising_risks") or []
        values = [
            r.id,
            r.get("software_function") or "—",
            r.get("software_item") or "—",
            r.get("hazard") or "[TODO]",
            r.get("initiating_causes") or "[TODO]",
            r.get("foreseeable_sequence") or "[TODO]",
            r.get("hazardous_situation") or "[TODO]",
            r.get("harm") or "[TODO]",
            prob or "—",
            sev or "—",
            risk_index(sev, prob) or "—",
            r.get("risk_level") or "—",
            "Yes" if r.get("acceptable") is True else ("No" if r.get("acceptable") is False else "—"),
            r.get("control_hierarchy") or "—",
            ", ".join(c.id for c in controls) if controls else "(none)",
            rprob or "—",
            rsev or "—",
            risk_index(rsev, rprob) or "—",
            r.get("residual_risk_level") or "—",
            "Yes" if r.get("residual_acceptable") is True else ("No" if r.get("residual_acceptable") is False else "—"),
            ", ".join(arising) if arising else "—",
            r.get("labeling_disclosure") or "n/a",
        ]
        # Residual Acceptable is column 20; map False to red.
        is_red = r.get("residual_acceptable") is False
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
            cell.alignment = WRAP_ALIGN
            if is_red and col_idx == 20:
                cell.fill = RED_FILL
                ctx.red_cells += 1


# ---------------------------------------------------------------------------
# Sheet 2: Production risk analysis (PRSK)
# ---------------------------------------------------------------------------


PROD_HEADERS = [
    "Risk ID", "Production Phase", "Asset at Risk", "Hazard",
    "Initiating causes", "Foreseeable sequence of events",
    "Hazardous situation", "Harm",
    "Initial Probability", "Initial Severity", "Initial Risk Index", "Initial Risk Level", "Initial Acceptable",
    "Control Hierarchy", "Mitigating Items",
    "Residual Probability", "Residual Severity", "Residual Risk Index", "Residual Risk Level", "Residual Acceptable",
]
PROD_WIDTHS = [16, 16, 24, 32, 40, 40, 32, 32, 14, 14, 12, 14, 14, 22, 22, 14, 14, 12, 14, 14]


def build_production_sheet(wb, ctx: BuildContext) -> None:
    ws = wb.create_sheet("Production risk analysis")
    apply_headers(ws, PROD_HEADERS, PROD_WIDTHS)
    active = [p for p in ctx.prsk if p.status != "Deprecated"]
    for row_idx, p in enumerate(sorted(active, key=lambda i: i.id), start=2):
        sev, prob = p.get("severity"), p.get("probability")
        rsev, rprob = p.get("residual_severity"), p.get("residual_probability")
        controls = ctx.controls_for(p.id)
        values = [
            p.id,
            p.get("production_phase") or "—",
            p.get("asset_at_risk") or "—",
            p.get("hazard") or "[TODO]",
            p.get("initiating_causes") or "[TODO]",
            p.get("foreseeable_sequence") or "[TODO]",
            p.get("hazardous_situation") or "[TODO]",
            p.get("harm") or "[TODO]",
            prob or "—",
            sev or "—",
            risk_index(sev, prob) or "—",
            p.get("risk_level") or "—",
            "Yes" if p.get("acceptable") is True else ("No" if p.get("acceptable") is False else "—"),
            p.get("control_hierarchy") or "—",
            ", ".join(c.id for c in controls) if controls else "(none)",
            rprob or "—",
            rsev or "—",
            risk_index(rsev, rprob) or "—",
            p.get("residual_risk_level") or "—",
            "Yes" if p.get("residual_acceptable") is True else ("No" if p.get("residual_acceptable") is False else "—"),
        ]
        is_red = p.get("residual_acceptable") is False
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
            cell.alignment = WRAP_ALIGN
            if is_red and col_idx == 20:
                cell.fill = RED_FILL
                ctx.red_cells += 1


# ---------------------------------------------------------------------------
# Sheet 3: Usability risk analysis (URSK)
# ---------------------------------------------------------------------------


URSK_HEADERS = [
    "Risk ID", "Use Scenario (USC)", "Use Error", "Hazard",
    "Hazardous Situation", "Harm",
    "Initial Likelihood", "Initial Severity", "Initial Risk Level", "Initial Acceptable",
    "Mitigating Items",
    "Residual Likelihood", "Residual Severity", "Residual Risk Level", "Residual Acceptable",
    "Triggered RSK (cascade)",
]
URSK_WIDTHS = [16, 18, 28, 32, 32, 32, 14, 14, 14, 14, 22, 14, 14, 14, 14, 22]


def build_usability_sheet(wb, ctx: BuildContext) -> None:
    ws = wb.create_sheet("Usability risk analysis")
    apply_headers(ws, URSK_HEADERS, URSK_WIDTHS)
    active = [u for u in ctx.ursk if u.status != "Deprecated"]
    for row_idx, u in enumerate(sorted(active, key=lambda i: i.id), start=2):
        controls = ctx.controls_for(u.id)
        triggers = (u.fm.get("links") or {}).get("triggers") or []
        values = [
            u.id,
            u.get("use_scenario") or "—",
            u.get("use_error") or "[TODO]",
            u.get("hazard") or "[TODO]",
            u.get("hazardous_situation") or "[TODO]",
            u.get("harm") or "[TODO]",
            u.get("likelihood") or "—",
            u.get("severity") or "—",
            u.get("risk_level") or "—",
            "Yes" if u.get("acceptable") is True else ("No" if u.get("acceptable") is False else "—"),
            ", ".join(c.id for c in controls) if controls else "(none)",
            u.get("residual_likelihood") or u.get("likelihood") or "—",
            u.get("residual_severity") or "—",
            u.get("residual_risk_level") or "—",
            "Yes" if u.get("residual_acceptable") is True else ("No" if u.get("residual_acceptable") is False else "—"),
            ", ".join(triggers) if triggers else "—",
        ]
        is_red = u.get("residual_acceptable") is False
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
            cell.alignment = WRAP_ALIGN
            if is_red and col_idx == 15:
                cell.fill = RED_FILL
                ctx.red_cells += 1


# ---------------------------------------------------------------------------
# Sheet 4: Cybersecurity risk analysis (THR)
# ---------------------------------------------------------------------------


CYBER_HEADERS = [
    "Risk ID", "STRIDE", "Attacker model", "Asset",
    "Vulnerability description",
    "Initial — Confidentiality", "Initial — Integrity", "Initial — Availability",
    "Mitigation / Remediation", "Mitigating Items",
    "Residual — Confidentiality", "Residual — Integrity", "Residual — Availability",
    "Residual Acceptable", "Triggered RSK (cascade)",
]
CYBER_WIDTHS = [16, 14, 18, 24, 40, 16, 16, 16, 32, 22, 16, 16, 16, 14, 22]


def build_cyber_sheet(wb, ctx: BuildContext) -> None:
    ws = wb.create_sheet("Cybersecurity risk analysis")
    apply_headers(ws, CYBER_HEADERS, CYBER_WIDTHS)
    active = [t for t in ctx.thr if t.status != "Deprecated"]
    for row_idx, t in enumerate(sorted(active, key=lambda i: i.id), start=2):
        controls = ctx.controls_for(t.id)
        triggers = (t.fm.get("links") or {}).get("triggers") or []
        stride = t.get("stride")
        stride_str = ", ".join(stride) if isinstance(stride, list) else str(stride or "—")
        # Vulnerability description = hazard if set, else title.
        vuln_desc = t.get("hazard") or t.title
        # Mitigation/remediation = informal "Expected controls" body — fallback to "(see linked items)"
        mitigation_text = "(see linked items)" if controls else "[TODO mitigation]"
        values = [
            t.id,
            stride_str,
            t.get("attacker") or "—",
            t.get("asset") or "—",
            vuln_desc,
            t.get("confidentiality_severity") or "n/a",
            t.get("integrity_severity") or "n/a",
            t.get("availability_severity") or "n/a",
            mitigation_text,
            ", ".join(c.id for c in controls) if controls else "(none)",
            t.get("residual_confidentiality_severity") or "n/a",
            t.get("residual_integrity_severity") or "n/a",
            t.get("residual_availability_severity") or "n/a",
            "Yes" if t.get("residual_acceptable") is True else ("No" if t.get("residual_acceptable") is False else "—"),
            ", ".join(triggers) if triggers else "—",
        ]
        is_red = t.get("residual_acceptable") is False
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
            cell.alignment = WRAP_ALIGN
            if is_red and col_idx == 14:
                cell.fill = RED_FILL
                ctx.red_cells += 1


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the QMS-ready 4-sheet Excel risk inventory.")
    parser.add_argument("--strict", action="store_true",
                        help="Exit 1 if any risk has residual_acceptable=False.")
    args = parser.parse_args()

    if not HAS_OPENPYXL:
        print(
            "ERROR: openpyxl is required but not installed.\n"
            "Install with:  pip install openpyxl  (or `uv pip install openpyxl`).\n"
            "The CSV inventory produced by /doc-risk-export remains available as a fallback.",
            file=sys.stderr,
        )
        return 1

    init_styles()

    config: dict = {}
    if CONFIG_PATH.is_file():
        try:
            config = parse_yaml(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"ERROR: failed to parse dt-config.yaml: {e}", file=sys.stderr)
            return 1
    else:
        print("WARN: dt-config.yaml not found — output filename will use 'UNKNOWN-V01'", file=sys.stderr)

    rsk = load_items("RSK", ITEMS_DIR)
    prsk = load_items("PRSK", ITEMS_DIR)
    ursk = load_items("URSK", ITEMS_DIR)
    thr = load_items("THR", ITEMS_DIR)

    if not (rsk or prsk or ursk or thr):
        print(
            "ERROR: no risk items found (RSK, PRSK, URSK, or THR). "
            "Run /doc-62304 first to populate docs/items/.",
            file=sys.stderr,
        )
        return 1

    mitigators: list[Item] = []
    for cat in ("SRS", "SDS", "TC"):
        mitigators += load_items(cat, ITEMS_DIR)

    ctx = BuildContext(
        config=config, rsk=rsk, prsk=prsk, ursk=ursk, thr=thr, mitigators=mitigators,
    )

    # Build workbook
    wb = Workbook()
    # Remove the default "Sheet" created by openpyxl
    default = wb.active
    wb.remove(default)

    build_design_sheet(wb, ctx)
    build_production_sheet(wb, ctx)
    build_usability_sheet(wb, ctx)
    build_cyber_sheet(wb, ctx)

    # Output paths
    doc = (config.get("document") or {}) if config else {}
    identifier = str(doc.get("identifier") or "UNKNOWN").strip()
    version_label = str(doc.get("version_label") or "V01").strip()
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = EXPORT_DIR / f"{identifier}-{version_label}-RISK-TABLE.xlsx"
    log_path = EXPORT_DIR / f"{identifier}-{version_label}-risk-xlsx.log"

    wb.save(xlsx_path)
    ctx.log(f"OK: wrote {xlsx_path.relative_to(ROOT)}")
    ctx.log(
        f"Items per sheet: Design={len(rsk)} · Production={len(prsk)} · "
        f"Usability={len(ursk)} · Cybersecurity={len(thr)}"
    )
    ctx.log(f"Red-highlighted cells (residual_acceptable=False): {ctx.red_cells}")

    # Log file
    header = [
        f"build_risk_xlsx run at {datetime.now(timezone.utc).isoformat(timespec='seconds')}",
        f"identifier={identifier} version_label={version_label}",
        "",
    ]
    log_path.write_text("\n".join(header + ctx.log_lines) + "\n", encoding="utf-8")
    print(str(xlsx_path))

    if args.strict and ctx.red_cells:
        print(f"STRICT: {ctx.red_cells} risks have residual_acceptable=False — failing", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
